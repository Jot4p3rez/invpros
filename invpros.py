import pandas as pd
import os
from datetime import datetime
import traceback
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


def process_inventories(config_usuario):
    log_mensajes = []
    try:
        wms_file_path = config_usuario["ruta_archivo_wms"]
        upc_sku_file_path = config_usuario["ruta_archivo_upc_sku"]
        erp_file_path = config_usuario["ruta_archivo_erp"]
        base_dir_salida = config_usuario["ruta_carpeta_salida"]
        wms_sheet_name = config_usuario.get("nombre_hoja_wms", "Sheet1")
        wms_col_item = config_usuario["col_item_wms"]
        wms_col_on_hand = config_usuario["col_on_hand_wms"]
        upc_sku_col_sku_para_wms = config_usuario["col_sku_en_upc_para_wms"]
        upc_sku_col_busr8 = config_usuario["col_busr8_upc_sku"]
        upc_sku_col_desc = config_usuario.get(
            "col_desc_upc_sku", "DESCR")
        erp_sheet_name = config_usuario.get("nombre_hoja_erp", "Sheet1")
        erp_col_key_for_busr8 = config_usuario["col_clave_erp_para_busr8"]
        erp_stock_on_hand_col_name = config_usuario["col_stock_erp"]
        erp_description_col_name = config_usuario.get(
            "col_descripcion_erp", "Description")
        wms_on_hand_final_col_name = "On Hand WMS"
        log_mensajes.append(f"Starting processing...")
        log_mensajes.append(f"Output directory: {base_dir_salida}")

        if not os.path.exists(wms_file_path):
            return f"Error: WMS file not found in '{wms_file_path}'"
        if not os.path.exists(upc_sku_file_path):
            return f"Error: UPC file not found in '{upc_sku_file_path}'"
        if not os.path.exists(erp_file_path):
            return f"Error: ERP file not found in '{erp_file_path}'"
        if not os.path.isdir(base_dir_salida):
            return f"Error: Invalid output folder: '{base_dir_salida}'"

        # --- Carga inicial de ERP y UPC para la validación de catálogo ---
        log_mensajes.append(
            f"\nLoading ERP and UPC for catalog validation (ERP.{erp_col_key_for_busr8} vs UPC.{upc_sku_col_busr8})...")

        dtype_erp_val = {erp_col_key_for_busr8: str}
        if erp_description_col_name:
            dtype_erp_val[erp_description_col_name] = str
        df_erp_val = pd.read_excel(
            erp_file_path, sheet_name=erp_sheet_name, dtype=dtype_erp_val)

        if erp_col_key_for_busr8 not in df_erp_val.columns:
            return f"Error: ERP key column '{erp_col_key_for_busr8}' not found in ERP file for validation."
        df_erp_val[erp_col_key_for_busr8] = df_erp_val[erp_col_key_for_busr8].astype(
            str).str.strip()

        # Preparar descripción ERP para matching posible
        erp_desc_col_exists = erp_description_col_name in df_erp_val.columns
        if erp_desc_col_exists:
            df_erp_val['erp_desc_normalized_short'] = df_erp_val[erp_description_col_name].fillna(
                "").astype(str).str.slice(0, 60).str.lower().str.strip()  # Cortar a 60 caracteres la descripcion
        else:
            log_mensajes.append(
                f"- WARNING: ERP description column '{erp_description_col_name}' not found. Possible match by description will be limited.")
            df_erp_val['erp_desc_normalized_short'] = ""

        log_mensajes.append(
            f"- ERP file for validation loaded (Column '{erp_col_key_for_busr8}' as string).")

        dtype_upc_sku_val = {
            upc_sku_col_busr8: str,
            upc_sku_col_sku_para_wms: str
        }
        if upc_sku_col_desc:
            dtype_upc_sku_val[upc_sku_col_desc] = str
        df_upc_sku_val = pd.read_excel(upc_sku_file_path, header=0, skiprows=[
                                       1], dtype=dtype_upc_sku_val)

        if upc_sku_col_busr8 not in df_upc_sku_val.columns:
            return f"Error: Column '{upc_sku_col_busr8}' not found in UPC for validation."
        df_upc_sku_val[upc_sku_col_busr8] = df_upc_sku_val[upc_sku_col_busr8].astype(
            str).str.strip()

        upc_desc_col_exists = upc_sku_col_desc in df_upc_sku_val.columns
        if upc_desc_col_exists:
            df_upc_sku_val['upc_desc_normalized'] = df_upc_sku_val[upc_sku_col_desc].fillna(
                "").astype(str).str.lower().str.strip()
        else:
            log_mensajes.append(
                f"- WARNING: UPC description column '{upc_sku_col_desc}' not found. Possible match by description will be limited.")
            df_upc_sku_val['upc_desc_normalized'] = ""

        df_maestro_busr8_unicos = df_upc_sku_val[[upc_sku_col_busr8]].copy(
        ).drop_duplicates(subset=[upc_sku_col_busr8])
        df_maestro_busr8_unicos.dropna(
            subset=[upc_sku_col_busr8], inplace=True)
        df_maestro_busr8_unicos = df_maestro_busr8_unicos[
            df_maestro_busr8_unicos[upc_sku_col_busr8] != '']
        log_mensajes.append(
            f"- UPC file for validation loaded and column '{upc_sku_col_busr8}' ready.")

        df_erp_vs_upc_merged = pd.merge(
            df_erp_val, df_maestro_busr8_unicos,
            left_on=erp_col_key_for_busr8, right_on=upc_sku_col_busr8,
            how='left', indicator=True
        )
        df_erp_sin_match_busr8_upc = df_erp_vs_upc_merged[df_erp_vs_upc_merged['_merge'] == 'left_only'].copy(
        )

        cols_to_drop_from_no_match = ['_merge']
        if upc_sku_col_busr8 in df_erp_sin_match_busr8_upc.columns:
            cols_to_drop_from_no_match.append(upc_sku_col_busr8)
        df_erp_sin_match_busr8_upc.drop(
            columns=cols_to_drop_from_no_match, inplace=True, errors='ignore')

        log_mensajes.append(
            f"- Found {len(df_erp_sin_match_busr8_upc)} ERP SKUs ('{erp_col_key_for_busr8}') with no direct match in UPC-SKU.BUSR8.")
        if erp_desc_col_exists and upc_desc_col_exists and \
           upc_sku_col_sku_para_wms in df_upc_sku_val.columns:
            log_mensajes.append(
                f"- Attempting to find possible matches by description for {len(df_erp_sin_match_busr8_upc)} unmatched ERP SKUs...")
            possible_matches_skus = []
            upc_desc_to_item_sku_map = {}
            if not df_upc_sku_val.empty:
                temp_upc_for_map = df_upc_sku_val[df_upc_sku_val['upc_desc_normalized'] != ""].copy(
                )
                temp_upc_for_map.drop_duplicates(
                    subset=['upc_desc_normalized'], keep='first', inplace=True)
                upc_desc_to_item_sku_map = temp_upc_for_map.set_index(
                    'upc_desc_normalized')[upc_sku_col_sku_para_wms].to_dict()

            for index, row in df_erp_sin_match_busr8_upc.iterrows():
                erp_desc_short_norm = row['erp_desc_normalized_short']
                if erp_desc_short_norm and erp_desc_short_norm in upc_desc_to_item_sku_map:
                    possible_matches_skus.append(
                        upc_desc_to_item_sku_map[erp_desc_short_norm])
                else:
                    possible_matches_skus.append("")
            if possible_matches_skus:
                df_erp_sin_match_busr8_upc['Possible Match SKU (from UPC)'] = possible_matches_skus
            else:
                df_erp_sin_match_busr8_upc['Possible Match SKU (from UPC)'] = ""

            log_mensajes.append(
                f"- 'Possible Match SKU (from UPC)' column added.")
        else:
            df_erp_sin_match_busr8_upc['Possible Match SKU (from UPC)'] = ""
            log_mensajes.append(
                f"- Skipped finding possible matches by description (required ERP or UPC description columns not found/specified).")

        if 'erp_desc_normalized_short' in df_erp_sin_match_busr8_upc.columns:
            df_erp_sin_match_busr8_upc.drop(
                columns=['erp_desc_normalized_short'], inplace=True, errors='ignore')

        # --- PASO 1 y 2: Cargar y fusionar WMS y UPC-SKU (para el flujo de inventario) ---
        log_mensajes.append(
            f"\nProcessing '{os.path.basename(wms_file_path)}' and '{os.path.basename(upc_sku_file_path)}' for inventory...")
        dtype_wms = {wms_col_item: str}
        df_wms = pd.read_excel(
            wms_file_path, sheet_name=wms_sheet_name, dtype=dtype_wms)
        log_mensajes.append(f"- '{os.path.basename(wms_file_path)}' loaded.")

        if upc_sku_col_sku_para_wms not in df_upc_sku_val.columns:
            return f"Error: Column '{upc_sku_col_sku_para_wms}' not found in UPC for WMS flow."
        df_upc_sku_val[upc_sku_col_sku_para_wms] = df_upc_sku_val[upc_sku_col_sku_para_wms].astype(
            str).str.strip()

        if wms_col_item not in df_wms.columns:
            return f"Error: Col. '{wms_col_item}' not in WMS."
        df_wms[wms_col_item] = df_wms[wms_col_item].astype(str).str.strip()

        df_upc_sku_seleccion_wms = df_upc_sku_val[[
            upc_sku_col_sku_para_wms, upc_sku_col_busr8]].copy()
        df_upc_sku_seleccion_wms.dropna(
            subset=[upc_sku_col_busr8, upc_sku_col_sku_para_wms], how='all', inplace=True)
        df_upc_sku_seleccion_wms = df_upc_sku_seleccion_wms[
            (df_upc_sku_seleccion_wms[upc_sku_col_busr8] != '') & (
                df_upc_sku_seleccion_wms[upc_sku_col_sku_para_wms] != '')
        ]
        df_upc_sku_seleccion_wms.drop_duplicates(
            subset=[upc_sku_col_sku_para_wms], keep='first', inplace=True)

        df_wms_consolidado = pd.merge(df_wms, df_upc_sku_seleccion_wms,
                                      left_on=wms_col_item, right_on=upc_sku_col_sku_para_wms, how='left')

        if upc_sku_col_sku_para_wms in df_wms_consolidado.columns and upc_sku_col_sku_para_wms != wms_col_item:
            df_wms_consolidado.drop(
                columns=[upc_sku_col_sku_para_wms], inplace=True)

        if upc_sku_col_busr8 in df_wms_consolidado.columns:
            df_wms_consolidado[upc_sku_col_busr8] = df_wms_consolidado[upc_sku_col_busr8].astype(
                str).str.strip()
        else:
            log_mensajes.append(
                f"WARNING: '{upc_sku_col_busr8}' not in WMS consolidated after merge. Creating empty column for '{upc_sku_col_busr8}'.")
            df_wms_consolidado[upc_sku_col_busr8] = ""
        log_mensajes.append(
            f"- UPC Data ('{upc_sku_col_busr8}') added to WMS.")

        # --- PASO 3: Crear Tabla Dinámica ---
        log_mensajes.append("\nCreating pivot table...")
        required_cols_pivot = [
            upc_sku_col_busr8, wms_col_on_hand]
        for col in required_cols_pivot:
            if col not in df_wms_consolidado.columns:
                return f"Error: Col '{col}' for pivot not in consolidated WMS ({df_wms_consolidado.columns.tolist()})."
        if wms_col_on_hand not in df_wms_consolidado.columns:
            return f"Error: Col '{wms_col_on_hand}' not in consolidated WMS."
        df_wms_consolidado[upc_sku_col_busr8].fillna(
            "N/A_BUSR8", inplace=True)

        tabla_dinamica_wms = pd.pivot_table(df_wms_consolidado,
                                            index=[upc_sku_col_busr8],
                                            values=wms_col_on_hand,
                                            aggfunc='sum', fill_value=0).reset_index()
        log_mensajes.append("- Pivot table created (grouped by BUSR8).")

        # --- PASO 4: Cruzar con ERP y añadir cálculos ---
        log_mensajes.append(
            f"\nCrossing with '{os.path.basename(erp_file_path)}' (Sheet: '{erp_sheet_name}')...")
        df_erp_main_flow = df_erp_val.copy()

        df_wms_para_cruce = tabla_dinamica_wms.rename(
            columns={wms_col_on_hand: wms_on_hand_final_col_name}
        )

        df_reporte_final = pd.merge(df_erp_main_flow, df_wms_para_cruce,
                                    left_on=erp_col_key_for_busr8,
                                    right_on=upc_sku_col_busr8,
                                    how='left')

        if upc_sku_col_busr8 in df_reporte_final.columns and upc_sku_col_busr8 != erp_col_key_for_busr8:
            df_reporte_final.drop(columns=[upc_sku_col_busr8], inplace=True)

        df_reporte_final[wms_on_hand_final_col_name].fillna(0, inplace=True)
        log_mensajes.append(
            f"- Merge with ERP completed. '{wms_on_hand_final_col_name}' filled with 0 where no match.")

        if 'erp_desc_normalized_short' in df_reporte_final.columns:
            df_reporte_final.drop(
                columns=['erp_desc_normalized_short'], inplace=True, errors='ignore')
        log_mensajes.append(
            f"\nAdding 'Difference' and 'Situation' columns...")
        if erp_stock_on_hand_col_name not in df_reporte_final.columns:
            log_mensajes.append(
                f"WARNING: ERP stock column '{erp_stock_on_hand_col_name}' not in final report. 'Difference' and 'Situation' will have error values.")
            df_reporte_final['Difference'] = f"Error: Column '{erp_stock_on_hand_col_name}' not found"
            df_reporte_final['Situation'] = "Error"
        else:
            df_reporte_final[erp_stock_on_hand_col_name] = pd.to_numeric(
                df_reporte_final[erp_stock_on_hand_col_name], errors='coerce').fillna(0)
            log_mensajes.append(
                f"- Column '{erp_stock_on_hand_col_name}' from ERP converted to numeric for calculations.")
            df_reporte_final[wms_on_hand_final_col_name] = pd.to_numeric(
                df_reporte_final[wms_on_hand_final_col_name], errors='coerce').fillna(0)
            df_reporte_final['Difference'] = df_reporte_final[erp_stock_on_hand_col_name] - \
                df_reporte_final[wms_on_hand_final_col_name]
            log_mensajes.append("- 'Difference' column calculated.")
            df_reporte_final['Situation'] = "OK"
            df_reporte_final.loc[df_reporte_final['Difference']
                                 != 0, 'Situation'] = "Review"
            if df_reporte_final['Difference'].apply(lambda x: isinstance(x, str)).any():
                df_reporte_final.loc[df_reporte_final['Difference'].apply(
                    lambda x: isinstance(x, str)), 'Situation'] = "Error"
            log_mensajes.append("- 'Situation' column calculated.")
        df_revision_urgente = pd.DataFrame()
        if 'Difference' in df_reporte_final.columns and pd.api.types.is_numeric_dtype(df_reporte_final['Difference']):
            condicion_revision = (df_reporte_final['Difference'] > 50) | (
                df_reporte_final['Difference'] < -50)
            df_revision_urgente = df_reporte_final[condicion_revision].copy()
            log_mensajes.append(
                f"- {len(df_revision_urgente)} rows found for 'Urgent Review' sheet (Difference > 50 or < -50).")

        # --- CREAR HOJA DE RESUMEN (DASHBOARD) ---
        log_mensajes.append(f"\nCreating Summary Dashboard sheet...")
        total_filas_erp_procesadas = len(df_reporte_final)
        total_skus_unicos_erp_reporte = "N/A"
        if erp_col_key_for_busr8 in df_reporte_final.columns:
            total_skus_unicos_erp_reporte = df_reporte_final[erp_col_key_for_busr8].nunique(
            )
        items_ok_val = "N/A"
        items_review_val = "N/A"
        if 'Situation' in df_reporte_final.columns:
            items_ok_val = len(
                df_reporte_final[df_reporte_final['Situation'] == 'OK'])
            items_review_val = len(
                df_reporte_final[df_reporte_final['Situation'] == 'Review'])
            items_error_val = len(
                df_reporte_final[df_reporte_final['Situation'] == 'Error'])
            if items_error_val > 0:
                items_review_val += items_error_val

        total_unidades_erp_val = "N/A"
        if erp_stock_on_hand_col_name in df_reporte_final.columns and pd.api.types.is_numeric_dtype(df_reporte_final[erp_stock_on_hand_col_name]):
            total_unidades_erp_val = df_reporte_final[erp_stock_on_hand_col_name].sum(
            )
        total_unidades_wms_val = "N/A"
        if wms_on_hand_final_col_name in df_reporte_final.columns and pd.api.types.is_numeric_dtype(df_reporte_final[wms_on_hand_final_col_name]):
            total_unidades_wms_val = df_reporte_final[wms_on_hand_final_col_name].sum(
            )
        diferencia_neta_unidades_val = "N/A"
        erp_units_is_num = pd.api.types.is_number(total_unidades_erp_val)
        wms_units_is_num = pd.api.types.is_number(total_unidades_wms_val)
        if erp_units_is_num and wms_units_is_num:
            diferencia_neta_unidades_val = float(
                total_unidades_erp_val) - float(total_unidades_wms_val)
        else:
            diferencia_neta_unidades_val = "N/A (Unit totals not numeric or not found)"
        items_erp_sin_match_upc_count_val = len(df_erp_sin_match_busr8_upc)
        ordered_metrics = ["Total Rows Processed from ERP", f"Total Unique ERP SKUs ({erp_col_key_for_busr8}) in Report", "Rows with OK Match (Difference = 0)", "Total Rows for Review (Difference != 0 or Error)",
                           "Total Units according to ERP", "Total Units according to WMS (for ERP SKUs)", "Net Unit Difference (ERP - WMS)", f"ERP SKUs not Matched in UPC-SKU Master (via {upc_sku_col_busr8})"]
        ordered_values = [total_filas_erp_procesadas, total_skus_unicos_erp_reporte, items_ok_val, items_review_val,
                          total_unidades_erp_val, total_unidades_wms_val, diferencia_neta_unidades_val, items_erp_sin_match_upc_count_val]
        resumen_data = {'Metric': ordered_metrics, 'Value': ordered_values}
        df_resumen = pd.DataFrame(resumen_data)
        log_mensajes.append("- Summary dashboard data calculated.")

        # --- Consolidar Top 5 ---
        df_top_5_consolidado = pd.DataFrame()
        if 'Difference' in df_reporte_final.columns and pd.api.types.is_numeric_dtype(df_reporte_final['Difference']):
            cols_para_top_display = [erp_col_key_for_busr8, erp_description_col_name,
                                     erp_stock_on_hand_col_name, wms_on_hand_final_col_name, 'Difference', 'Situation']
            cols_presentes_para_top_display = [
                col for col in cols_para_top_display if col in df_reporte_final.columns]
            if len(cols_presentes_para_top_display) >= 2:
                df_reporte_final_sorted = df_reporte_final.sort_values(
                    by='Difference', ascending=False)
                df_top_positivas_calc = df_reporte_final_sorted[df_reporte_final_sorted['Difference'] > 0].head(5)[
                    cols_presentes_para_top_display]
                df_top_negativas_calc = df_reporte_final_sorted[df_reporte_final_sorted['Difference'] < 0].tail(
                    5)[cols_presentes_para_top_display].sort_values(by='Difference', ascending=True)
                if not df_top_positivas_calc.empty:
                    titulo_positivas = {col: ("--- Top 5 Positive Differences ---" if col ==
                                              cols_presentes_para_top_display[0] else "") for col in cols_presentes_para_top_display}
                    df_titulo_positivas = pd.DataFrame([titulo_positivas])
                    df_top_5_consolidado = pd.concat(
                        [df_top_5_consolidado, df_titulo_positivas, df_top_positivas_calc], ignore_index=True)
                if not df_top_negativas_calc.empty:
                    titulo_negativas = {col: ("--- Top 5 Negative Differences ---" if col ==
                                              cols_presentes_para_top_display[0] else "") for col in cols_presentes_para_top_display}
                    df_titulo_negativas = pd.DataFrame([titulo_negativas])
                    if not df_top_5_consolidado.empty and not df_top_positivas_calc.empty:
                        df_top_5_consolidado = pd.concat([df_top_5_consolidado, pd.DataFrame(
                            [{col: "" for col in cols_presentes_para_top_display}])], ignore_index=True)
                    df_top_5_consolidado = pd.concat(
                        [df_top_5_consolidado, df_titulo_negativas, df_top_negativas_calc], ignore_index=True)
                if not df_top_5_consolidado.empty:
                    log_mensajes.append("- Top 5 differences consolidated.")

        fecha_actual_dt = datetime.now()
        fecha_actual_str = fecha_actual_dt.strftime(config_usuario.get(
            "formato_fecha_salida", "%Y%m%d%H%M"))
        nombre_base_archivo_salida = config_usuario.get(
            "nombre_base_reporte", "InventoryCrossReportFinished")
        output_filename = os.path.join(
            base_dir_salida, f"{nombre_base_archivo_salida}_{fecha_actual_str}.xlsx")

        new_cols_order_final_applied = False
        new_cols_order = []
        try:
            cols_final = df_reporte_final.columns.tolist()
            difference_col_name = 'Difference'
            situation_col_name = 'Situation'
            if wms_on_hand_final_col_name in cols_final and difference_col_name in cols_final and situation_col_name in cols_final:
                cols_final_copy = cols_final[:]
                if difference_col_name in cols_final_copy:
                    cols_final_copy.pop(
                        cols_final_copy.index(difference_col_name))
                if situation_col_name in cols_final_copy:
                    cols_final_copy.pop(
                        cols_final_copy.index(situation_col_name))
                idx_on_hand_wms = cols_final_copy.index(
                    wms_on_hand_final_col_name)
                new_cols_order = cols_final_copy[:idx_on_hand_wms+1] + [
                    difference_col_name, situation_col_name] + cols_final_copy[idx_on_hand_wms+1:]
                df_reporte_final = df_reporte_final[new_cols_order]
                log_mensajes.append("- 'Full Report' columns reordered.")
                new_cols_order_final_applied = True
            else:
                log_mensajes.append(
                    "- Warning: Columns were not reordered in 'Full Report' (expected columns not found).")
        except (ValueError, KeyError) as e_reorder_final:
            log_mensajes.append(
                f"- Warning: Error when reordering columns in 'Full Report' ({e_reorder_final}).")

        if not df_revision_urgente.empty and new_cols_order_final_applied:
            try:
                if all(col in df_revision_urgente.columns for col in new_cols_order):
                    df_revision_urgente = df_revision_urgente[new_cols_order]
                    log_mensajes.append("- 'Urgent Review' columns reordered.")
                else:
                    log_mensajes.append(
                        "- Warning: Not all columns for reordering 'Urgent Review' were present in it.")
            except (ValueError, KeyError) as e_reorder_rev:
                log_mensajes.append(
                    f"- Warning: Error when reordering columns in 'Urgent Review' ({e_reorder_rev}).")

        with pd.ExcelWriter(output_filename, engine='openpyxl') as writer:
            df_reporte_final.to_excel(
                writer, sheet_name='Full Report', index=False)
            log_mensajes.append(f"- 'Full Report' sheet saved.")

            try:
                workbook = writer.book
                worksheet = writer.sheets['Full Report']
                red_font = Font(color="FF0000")
                header_row_values = [cell.value for cell in worksheet[1]]
                situation_col_letter = None
                difference_col_letter = None
                situation_col_name_in_excel = 'Situation'
                difference_col_name_in_excel = 'Difference'
                try:
                    situation_col_idx_excel = header_row_values.index(
                        situation_col_name_in_excel) + 1
                    situation_col_letter = get_column_letter(
                        situation_col_idx_excel)
                except ValueError:
                    log_mensajes.append(
                        f"- WARNING: Column '{situation_col_name_in_excel}' not found in 'Full Report' for formatting.")
                try:
                    difference_col_idx_excel = header_row_values.index(
                        difference_col_name_in_excel) + 1
                    difference_col_letter = get_column_letter(
                        difference_col_idx_excel)
                except ValueError:
                    log_mensajes.append(
                        f"- WARNING: Column '{difference_col_name_in_excel}' not found in 'Full Report' for formatting.")
                if situation_col_letter and difference_col_letter:
                    for row_idx in range(2, worksheet.max_row + 1):
                        situation_cell = worksheet[f"{situation_col_letter}{row_idx}"]
                        if situation_cell.value == "Review":
                            situation_cell.font = red_font
                            worksheet[f"{difference_col_letter}{row_idx}"].font = red_font
                    log_mensajes.append(
                        f"- Conditional formatting applied to 'Full Report' sheet.")
                else:
                    log_mensajes.append(
                        f"- WARNING: Could not apply conditional formatting. Required columns not found by header.")
            except Exception as e_format:
                log_mensajes.append(
                    f"- WARNING: Could not apply conditional formatting. Error: {e_format}")

            df_resumen.to_excel(
                writer, sheet_name='Summary Dashboard', index=False)
            log_mensajes.append(f"- 'Summary Dashboard' sheet saved.")

            if not df_top_5_consolidado.empty:
                df_top_5_consolidado.to_excel(
                    writer, sheet_name='Top 5 Difference', index=False)
                log_mensajes.append(f"- 'Top 5 Difference' sheet saved.")
            else:
                log_mensajes.append(
                    f"- 'Top 5 Difference' sheet not generated (no top differences found).")

            if not df_revision_urgente.empty:
                df_revision_urgente.to_excel(
                    writer, sheet_name='Urgent Review (Diff > 50)', index=False)
                log_mensajes.append(
                    f"- 'Urgent Review (Diff > 50)' sheet saved.")
            else:
                log_mensajes.append(
                    "- 'Urgent Review (Diff > 50)' sheet not generated (no rows met criteria).")

            if not df_erp_sin_match_busr8_upc.empty:

                df_erp_sin_match_busr8_upc.to_excel(
                    writer, sheet_name='ERP Sku no Match in UPC BUSR8', index=False)
                log_mensajes.append(
                    f"- 'ERP Sku no Match in UPC BUSR8' sheet saved.")
            else:
                log_mensajes.append(
                    "- 'ERP Sku no Match in UPC BUSR8' sheet not generated (all ERP Skus matched UPC BUSR8).")

        log_mensajes.append(
            f"\nProcess completed! The final report with multiple sheets has been saved as: '{output_filename}'")
        return "\n".join(log_mensajes)

    except FileNotFoundError as e:
        log_mensajes.append(f"Critical File Not Found Error: {e}.")
        return "\n".join(log_mensajes)
    except KeyError as e:
        log_mensajes.append(
            f"Critical Column Name Error (KeyError): {e}. Check column configuration in the GUI and file headers.")
        return "\n".join(log_mensajes)
    except Exception as e:
        log_mensajes.append(f"An unexpected error occurred:\n{e}")
        log_mensajes.append("\nTechnical details (stack trace):")
        log_mensajes.append(traceback.format_exc())
        return "\n".join(log_mensajes)
